import openpyxl
#jhale changes

class DataReader:
    @staticmethod
    def readDataforgrpsave():
        path = "D:\demo.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook["Sheet4"]
        rows = sheet.max_row
        cols = sheet.max_column
        tcId = []
        userName = []
        passWord = []
        groupName = []
        permission = []
        cnt=0
        print(rows)

        for i in range(2, rows + 1):
            tcId.append(sheet.cell(row=i, column=1).value)
            userName.append(sheet.cell(row=i, column=2).value)
            passWord.append(sheet.cell(row=i, column=3).value)
            groupName.append(sheet.cell(row=i, column=4).value)
            permission.append(sheet.cell(row=i, column=5).value)
            x=rows-1
            print(x)
        return tcId,userName,passWord,groupName,permission,x,cnt

#
ob=DataReader
ob.readDataforgrpsave()

# if __name__=="__main__":
#     DataReader.readDataforlogin()
